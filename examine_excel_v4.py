import openpyxl
file_path = "/Users/josh/Desktop/CotizadorAdeslas/Cotizador Base 2026 2.0  (1).xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['Cotizador']

# Check row 4 or 5 for labels
labels = []
for c in range(14, 25): # N to V
    val = sheet.cell(row=5, column=c).value
    if not val:
        val = sheet.cell(row=4, column=c).value
    labels.append(f"{sheet.cell(row=5, column=c).coordinate}: {val}")

print("Column labels:", labels)
# Check Row 7 as well for more info
print("Sub-labels (row 6):", [sheet.cell(row=6, column=c).value for c in range(14, 25)])
