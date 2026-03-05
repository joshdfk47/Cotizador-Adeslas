import openpyxl
file_path = "/Users/josh/Desktop/CotizadorAdeslas/Cotizador Base 2026 2.0  (1).xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
wb_f = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Cotizador']
sheet_f = wb_f['Cotizador']

for r in range(1, 40):
    val = sheet.cell(row=r, column=13).value # M column (13) where labels are
    if val and "VALOR DE VENTA" in val.upper():
        print(f"Found '{val}' at {sheet.cell(row=r, column=13).coordinate}")
        for c in range(14, 25):
            print(f"  {sheet.cell(row=r, column=c).coordinate}: {sheet.cell(row=r, column=c).value} ({sheet_f.cell(row=r, column=c).value})")
