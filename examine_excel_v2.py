import openpyxl
file_path = "/Users/josh/Desktop/CotizadorAdeslas/Cotizador Base 2026 2.0  (1).xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
wb_f = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['Cotizador']
sheet_f = wb_f['Cotizador']

for r in range(1, 50):
    for c in range(1, 20):
        val = sheet.cell(row=r, column=c).value
        if val and isinstance(val, str) and "PRECIOS MEDIOS" in val.upper():
            print(f"Found '{val}' at {sheet.cell(row=r, column=c).coordinate}")
            # Look at the values below
            for r_off in range(1, 5):
                print(f"  Row {r+r_off}: ", end="")
                for c_off in range(0, 10):
                    cv = sheet.cell(row=r+r_off, column=c+c_off).value
                    cf = sheet_f.cell(row=r+r_off, column=c+c_off).value
                    print(f"{cv} ({cf}) | ", end="")
                print()
