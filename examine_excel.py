import openpyxl
import os

file_path = "/Users/josh/Desktop/CotizadorAdeslas/Cotizador Base 2026 2.0  (1).xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    # Also load with formulas
    wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_f = wb_formulas[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        
        # Search for interesting keywords
        for row in range(1, min(sheet.max_row, 100)):
            for col in range(1, min(sheet.max_column, 20)):
                cell = sheet.cell(row=row, column=col)
                value = cell.value
                if value and isinstance(value, str):
                    low_val = value.lower()
                    if "prima anual" in low_val or "precio medio" in low_val or "promocion" in low_val or "promoción" in low_val:
                        formula = sheet_f.cell(row=row, column=col).value
                        print(f"Found '{value}' at {cell.coordinate}. Formula: {formula}")
                        
                        # Print some surrounding cells to understand context
                        for r_off in range(0, 3):
                            row_cells = []
                            for c_off in range(1, 10):
                                c = sheet.cell(row=row + r_off, column=col + c_off)
                                cf = sheet_f.cell(row=row + r_off, column=col + c_off)
                                row_cells.append(f"{c.coordinate}: {c.value} ({cf.value})")
                            print(f"  Row {row+r_off}: {' | '.join(row_cells)}")

except Exception as e:
    print(f"Error: {e}")
