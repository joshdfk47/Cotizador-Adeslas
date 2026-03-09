import openpyxl

wb = openpyxl.load_workbook('Cotizador Base 2026 2.0  (2).xlsx', data_only=False)
ws_c = wb['Cotizador']

# Column for Seniors is U (21), for Plena is Q (17)
print("Plena (17):", ws_c.cell(row=39, column=17).value)
print("Seniors (21):", ws_c.cell(row=39, column=21).value)
