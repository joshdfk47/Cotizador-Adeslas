import pandas as pd
import json
import re
import openpyxl

EXCEL_FILE = 'Cotizador Base 2026 2.0  (2).xlsx'
OUTPUT_FILE = 'data_autogen.js'

def safe_float(val, default=0.0):
    try:
        if pd.isna(val): return default
        return float(val)
    except: return default

# ───────────────────────────────────────────────────────
# 1. EXTRAER TABLA DE BÚSQUEDA (Precios: Col A-D)
# ───────────────────────────────────────────────────────
def extract_prices(ws):
    price_table = {}  # key: "Product Name Zone"
    for r in range(2, ws.max_row + 1):
        e_min = ws.cell(row=r, column=1).value
        e_max = ws.cell(row=r, column=2).value
        p_name = ws.cell(row=r, column=3).value
        p_val = ws.cell(row=r, column=4).value
        
        if p_name is None: continue
        p_name = str(p_name).strip()
        
        if p_name not in price_table:
            price_table[p_name] = []
            
        if isinstance(p_val, str) and p_val.lower().strip() == "no asegurable":
            parsed_p = "No asegurable"
        else:
            try:
                parsed_p = round(float(p_val), 2) if p_val is not None else 0.0
            except:
                parsed_p = 0.0

        price_table[p_name].append({
            "min": int(float(e_min)) if e_min is not None else 0,
            "max": int(float(e_max)) if e_max is not None else 120,
            "price": parsed_p
        })
    return price_table

# ───────────────────────────────────────────────────────
# 2. DENTAL (Precios: Col F-H)
# ───────────────────────────────────────────────────────
def extract_dental(ws):
    d_pp = {}
    d_total = {}
    for r in range(2, 20):
        n  = ws.cell(row=r, column=6).value
        pp = ws.cell(row=r, column=7).value
        tot = ws.cell(row=r, column=8).value
        if n is None:
            break
        k = str(int(float(n)))
        d_pp[k]    = round(float(pp), 4) if pp else 0.0
        d_total[k] = round(float(tot), 4) if tot else 0.0
    return d_pp, d_total

# ───────────────────────────────────────────────────────
# 3. MAPA CP → ZONA (AuxCP)
# ───────────────────────────────────────────────────────
def extract_cp_map(ws):
    cp_map = {}
    for r in range(2, 60):
        prov = ws.cell(row=r, column=1).value
        cp_pre = ws.cell(row=r, column=2).value
        zona = ws.cell(row=r, column=4).value
        if cp_pre is None:
            continue
        prefix = str(int(float(cp_pre))).zfill(2)
        cp_map[prefix] = {
            'provincia': str(prov),
            'zona': int(float(zona)) if zona else 1
        }
    return cp_map

# ───────────────────────────────────────────────────────
# 4. EXCEPCIONES
# ───────────────────────────────────────────────────────
def extract_excepciones(ws):
    exc = {}
    for r in range(2, 30):
        n = ws.cell(row=r, column=1).value
        v = ws.cell(row=r, column=2).value
        if n is None:
            break
        exc[str(n).strip()] = float(v) if v is not None else 0.0
    return exc

def build_product_rules():
    graduated_total = {
        "graduated": [
            {"min_count": 5, "factor": 0.85},
            {"min_count": 4, "factor": 0.9},
            {"min_count": 3, "factor": 0.95}
        ]
    }
    threshold_4 = {"threshold": 4, "factor": 0.9}
    threshold_2 = {"threshold": 2, "factor": 0.9}
    threshold_3 = {"threshold": 3, "factor": 0.9}

    return {
        "Go": {
            "base_name": "Go",
            "multi_discount": threshold_2,
            "k6_applies": True,
            "k8_applies": False,
            "dental_mode": "per_person_inherited",
            "age_min": 0, "age_max": 120,
            "pensionista_discount": False,
        },
        "Plena Vital": {
            "base_name": "Plena Vital",
            "multi_discount": threshold_4,
            "k6_applies": True,
            "k8_applies": True,
            "dental_mode": "per_person_inherited",
            "age_min": 0, "age_max": 120,
            "pensionista_discount": True,
        },
        "Adeslas Plena": {
            "base_name": "Plena",
            "multi_discount": threshold_4,
            "k6_applies": True,
            "k8_applies": True,
            "dental_mode": "per_person_inherited",
            "age_min": 0, "age_max": 120,
            "pensionista_discount": True,
        },
        "Adeslas Plena Plus": {
            "base_name": "Plena Plus",
            "multi_discount": threshold_4,
            "k6_applies": True,
            "k8_applies": True,
            "dental_mode": "per_person_inherited",
            "age_min": 0, "age_max": 120,
            "pensionista_discount": True,
        },
        "Adeslas Plena Extra 150": {
            "base_name": "Plena Extra 150",
            "multi_discount": threshold_4,
            "k6_applies": True,
            "k8_applies": True,
            "dental_mode": "per_person_direct",
            "age_min": 0, "age_max": 120,
            "pensionista_discount": True,
        },
        "Plena Total Vital": {
            "base_name": "Plena Total Vital",
            "multi_discount": graduated_total,
            "k6_applies": False,
            "k8_applies": True,
            "dental_mode": "none",
            "only_con_dental": True,
            "age_min": 0, "age_max": 120,
            "payment": "monthly_only",
            "pensionista_excluded": True,
        },
        "Adeslas Plena Total": {
            "base_name": "Plena Total",
            "multi_discount": graduated_total,
            "k6_applies": False,
            "k8_applies": True,
            "dental_mode": "none",
            "only_con_dental": True,
            "age_min": 0, "age_max": 120,
            "payment": "monthly_only",
            "pensionista_excluded": True,
        },
        "Adeslas Pymes TOTAL": {
            "base_name": "Pymes Total",
            "multi_discount": None,
            "k8_applies": "double",
            "dental_mode": "none",
            "only_con_dental": True,
            "age_min": 0, "age_max": 120,
            "pensionista_excluded": True,
        },
        "Adeslas Seniors": {
            "base_name": "Seniors",
            "multi_discount": threshold_3,
            "multi_discount_total": threshold_4,
            "k6_applies": False,
            "k8_applies": True,
            "dental_mode": "per_person_inherited",
            "age_min": 55, "age_max": 120,
            "pensionista_discount": True,
            "pensionista_excluded": False,
        },
        "Adeslas Plena Total Seniors": {
            "base_name": "Seniors total",
            "multi_discount": None,
            "k6_applies": False,
            "k8_applies": True,
            "dental_mode": "none",
            "only_con_dental": True,
            "age_min": 60, "age_max": 120,
            "payment": "monthly_only",
            "pensionista_excluded": True,
        },
        "Adeslas NIF": {
            "base_name": "Negocios NIF",
            "multi_discount": None,
            "k6_applies": False,
            "k8_applies": False,
            "dental_mode": "total_at_end",
            "age_min": 0, "age_max": 120,
            "pensionista_excluded": True,
        },
    }

def main():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    price_table = extract_prices(wb['Precios'])
    d_pp, d_total = extract_dental(wb['Precios'])
    cp_map = extract_cp_map(wb['AuxCP'])
    exc = extract_excepciones(wb['Excepciones'])
    
    data = {
        "products": sorted(list(build_product_rules().keys())),
        "price_table": price_table,
        "dental_per_person": d_pp,
        "dental_total": d_total,
        "cp_map": cp_map,
        "product_rules": build_product_rules(),
        "excepciones": exc
    }
    
    with open(OUTPUT_FILE, 'w') as f:
        f.write("const DATA = " + json.dumps(data, indent=2) + ";")
    
    print(f"Generado {OUTPUT_FILE}")
    print(f"  Productos: {len(data['products'])}")
    print(f"  Tablas de precios: {len(price_table)} bases")
    print(f"  Mapa CP: {len(cp_map)} prefijos")

if __name__ == "__main__":
    main()
